"""

"""
import os,sys
import numpy as np
import pandas as pd
import scipy.stats as stats
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from cartopy.mpl.ticker import LongitudeFormatter, LatitudeFormatter
import matplotlib.ticker as mticker
import matplotlib.patheffects as PathEffects
import matplotlib.pyplot as plt
import pyskew.plot_skewness as psk
import pyskew.utilities as utl
import pyskew.plot_gravity as pg
from geographiclib.geodesic import Geodesic
import pyrot.apwp as apwp
from pyrot.rot import Rot,get_pole_arc_misfit_uncertainty,fast_fit_circ,plot_error_from_points
import pyrot.max as pymax
from multi_circ_inv import fit_circ,pole_arc_fitfunc
from functools import reduce
from time import time
from scipy.interpolate import PchipInterpolator
import pmagpy.pmag as pmag
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import load_workbook

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    return writer

fin = sys.argv[1]
inv_data = pd.read_csv(fin,index_col=0,dtype={"BendLat":float,"BendLon":float,"EMDis":float,"EMLon":float,"EMLat":float,"EMErr":float,"EM_Start_Dis":float,"EM_start_azi":float,"EMpols":object,"EMsds":object,"HIDis":float,"HILon":float,"HILat":float,"HIErr":float,"HI_Start_Dis":float,"HI_start_azi":float,"HIpols":object,"HIsds":object}).T
for hs in inv_data.columns:
    try:
        inv_data[hs]["HIpols"] = list(map(float,inv_data[hs]["HIpols"].strip("[ ]").split()))
        inv_data[hs]["HIsds"] = list(map(float,inv_data[hs]["HIsds"].strip("[ ]").split()))
        inv_data[hs]["EMpols"] = list(map(float,inv_data[hs]["EMpols"].strip("[ ]").split()))
        inv_data[hs]["EMsds"] = list(map(float,inv_data[hs]["EMsds"].strip("[ ]").split()))
    except ValueError:
        inv_data[hs]["HIpols"] = list(map(float,inv_data[hs]["HIpols"].strip("[ ]").split(",")))
        inv_data[hs]["HIsds"] = list(map(float,inv_data[hs]["HIsds"].strip("[ ]").split(",")))
        inv_data[hs]["EMpols"] = list(map(float,inv_data[hs]["EMpols"].strip("[ ]").split(",")))
        inv_data[hs]["EMsds"] = list(map(float,inv_data[hs]["EMsds"].strip("[ ]").split(",")))
padding = .05
geoid = Geodesic(6371.,0.)
hi_color = "#C4A58E"#plt.rcParams['axes.prop_cycle'].by_key()['color'][2]
em_color = plt.rcParams['axes.prop_cycle'].by_key()['color'][1]
tr_color = plt.rcParams['axes.prop_cycle'].by_key()['color'][0]
undated_color = plt.rcParams['axes.prop_cycle'].by_key()['color'][3]
mean_color = "k"#plt.rcParams['axes.prop_cycle'].by_key()['color'][2]
land_resolution = "10m"
decimation = 16
fout = os.path.join("./data",os.path.basename(fin).split(".")[0] + "_recdata.xlsx")
hotspot_list = ["Hawaii","Rurutu","Louisville (Heaton & Koppers 2019)"]
min_age,max_age,age_step = 0.,80.,1.
rurutu_gap = [10.,48.]
rurutu_max_age = 80.
ages = np.arange(min_age,max_age+age_step,age_step)
reconstruction_anomalies_df = pd.DataFrame([range(len(ages)-1),ages[1:]],index=["Name","Age"]).T
f = append_df_to_excel(fout, reconstruction_anomalies_df, sheet_name="Magnetic Anomalies", index=False, startrow=0)
f.save(); f.close()

bfig = plt.figure(figsize=(6,18),dpi=200)
bax_pos = 311
for k,hs1 in enumerate(hotspot_list):
    print(hs1)
    hs_df = pd.DataFrame()
    color = plt.rcParams['axes.prop_cycle'].by_key()['color'][k+2]
    if "subair" in sys.argv[1]: data = pd.read_excel("data/pa_seamount_ages_subareal_included.xlsx",hs1)
    else: data = pd.read_excel("data/pa_seamount_ages_updated_Koopers2019.xlsx",hs1)
    data = data[data["Quality"]=="g"]

    if hs1 == "Hawaii":
        hi_data = data[data["Latitude"]<33.]
        tr_data = data[(data["Latitude"]>33.) & (data["Latitude"]<33.)]
        em_data = data[data["Latitude"]>33.]
    elif hs1 == "Louisville" or hs1 == "Louisville (Heaton & Koppers 2019)":
        hi_data = data[data["Longitude"]>-169.]
        tr_data = data[(data["Longitude"]>-169.) & (data["Longitude"]<=-169.)]
        em_data = data[data["Longitude"]<-169.]
    elif hs1 == "Rurutu":
        cut_age = 20.
        hi_data = data[data["Age (Ma)"]<cut_age]
        tr_data = data[(data["Age (Ma)"]>cut_age) & (data["Age (Ma)"]<cut_age)]
        em_data = data[data["Age (Ma)"]>=cut_age]
    else: raise IOError("No HS Track named %s known must edit script"%hs1)
    dated_hi_data = hi_data[hi_data["Age (Ma)"].notnull()]

    hi_start_azi = inv_data[hs1]["HI_start_azi"]
    for i,datum in dated_hi_data.iterrows():
        geodict = geoid.Inverse(datum["Latitude"],datum["Longitude"],inv_data[hs1]["HILat"],inv_data[hs1]["HILon"])
        dis = (geodict["azi2"]-hi_start_azi)*np.sin(np.deg2rad(inv_data[hs1]["HIDis"]))
        dated_hi_data.at[i,"Dis"] = dis
    dated_hi_data.sort_values("Dis",inplace=True)

    good_dated_hi_data = dated_hi_data[dated_hi_data["Quality"]=="g"]

    if hs1 == "Hawaii": bend_extent = [150.,210.,13.,55.]
    elif hs1 == "Louisville" or hs1 == "Louisville (Heaton & Koppers 2019)": bend_extent = [175.,225.,-57.,-20.]
    elif hs1 == "Rurutu": bend_extent = [160.,215.,-30.,10.]
    else: raise IOError("No HS Track named %s known must edit script"%hs1)

    bproj = ccrs.Mercator(central_longitude=180)
    bm = bfig.add_subplot(bax_pos,projection=bproj)
    bm.set_xticks(np.arange(0, 370, 10.), crs=ccrs.PlateCarree())
    bm.set_yticks(np.arange(-80, 90, 10.), crs=ccrs.PlateCarree())
    bm.tick_params(grid_linewidth=1.0,grid_linestyle=":",color="grey",labelsize=6,tickdir="in",left=True,top=True)
    lon_formatter = LongitudeFormatter(zero_direction_label=True)
    lat_formatter = LatitudeFormatter()
    bm.xaxis.set_major_formatter(lon_formatter)
    bm.yaxis.set_major_formatter(lat_formatter)
    bm.outline_patch.set_linewidth(0.5)
    bm.coastlines(linewidth=2,color="k",resolution=land_resolution)
    bm.annotate(chr(64+bax_pos%10)+")",xy=(1-0.04,0.04),xycoords="axes fraction",bbox=dict(boxstyle="round", fc="w",alpha=.7),va="bottom",ha="right")

    if hs1=="Rurutu":
        ages = np.array(list(np.arange(min_age,rurutu_gap[0]+age_step,age_step))+list(np.arange(rurutu_gap[1],rurutu_max_age+age_step,age_step)))
    else:
        ages = np.arange(min_age,max_age+age_step,age_step)

    for idx,midpoint_age in enumerate(ages):
        if midpoint_age<inv_data[hs1]["BendAge"]:  ##########Hawaiian
            hs1_b = np.sqrt(2)*(33./111.113)*np.sqrt(inv_data[hs1]["HIErr"]/(len(data.index)-6))
            if len(inv_data[hs1]["HIpols"])==2:
                midpoint_dis = ((midpoint_age-inv_data[hs1]["HIpols"][1])/inv_data[hs1]["HIpols"][0] - inv_data[hs1]["HI_Start_Dis"])/np.sin(np.deg2rad(inv_data[hs1]["HIDis"]))
                hs1_a = np.sqrt(2*(((inv_data[hs1]["HIsds"][1]**2)*(1/inv_data[hs1]["HIpols"][0])**2) + ((inv_data[hs1]["HIsds"][0]**2)*(-(midpoint_age-inv_data[hs1]["HI_mean_age"]-inv_data[hs1]["HIpols"][1])/(inv_data[hs1]["HIpols"][0]**2))**2)) + hs1_b**2)
            elif len(inv_data[hs1]["HIpols"])==3:
                midpoint_dis = ((((-inv_data[hs1]["HIpols"][1]+np.sqrt(inv_data[hs1]["HIpols"][1]**2 - 4*inv_data[hs1]["HIpols"][0]*(inv_data[hs1]["HIpols"][2]-midpoint_age)))/(2*inv_data[hs1]["HIpols"][0]))) - inv_data[hs1]["HI_Start_Dis"])/np.sin(np.deg2rad(inv_data[hs1]["HIDis"]))
#                x = np.linspace(-1,inv_data[hs1]["EM_Start_Dis"]/np.sin(np.deg2rad(inv_data[hs1]["HIDis"])),10000)
#                idx = np.argwhere(np.diff(np.sign(x-midpoint_dis))).flatten()[0]
#                age_err = (utl.polyenv(inv_data[hs1]["HIpols"],x,np.ones(len(x))*inv_data[hs1]["HIAgeSd"],center=inv_data[hs1]["HI_mean_dis"])/np.sqrt(inv_data[hs1]["HI_dated_N"]))[idx]
#                rate = inv_data[hs1]["HIpols"][0]*midpoint_dis + inv_data[hs1]["HIpols"][1]
#                hs1_a = np.sqrt((age_err*rate)**2 + hs1_b**2)
#                print(midpoint_age,midpoint_dis,idx,age_err,111.113*rate,hs1_a)
                x = np.linspace(0,inv_data[hs1]["EM_Start_Dis"],1000)
#                center_pols,center_sds = utl.polyrecenter(inv_data[hs1]["HIpols"],x,np.ones(len(x))*inv_data[hs1]["HIAgeSd"],center=inv_data[hs1]["HI_mean_dis"])
                center_pols,center_sds = inv_data[hs1]["HIpols"],inv_data[hs1]["HIsds"]
                psi = np.sqrt(center_pols[1]**2 - 4*center_pols[0]*(center_pols[2]-(midpoint_age)))
                d_dc = -1/psi
                d_db = (center_pols[1]/psi-1)/(2*center_pols[0])
                d_da = -(psi-center_pols[1])/(2*center_pols[0]**2) - (center_pols[2]-(midpoint_age))/(center_pols[0]*psi)
                hs1_a = np.sqrt(2*(((center_sds[2]**2)*(d_dc**2)) + ((center_sds[1]**2)*(d_db**2)) + ((center_sds[0]**2)*(d_da**2))) + hs1_b**2)
#                import pdb; pdb.set_trace()
#                hs1_a = np.sqrt(2)*np.sqrt((((center_sds[2]**2)*(d_dc**2)) + ((center_sds[1]**2)*(d_db**2)) + ((center_sds[0]**2)*(d_da**2)))/(len(x)*np.sin(np.deg2rad(inv_data[hs1]["HIDis"]))**2) + hs1_b**2)
            else: raise ValueError("Degree %d inverse not programmed yet"%(len(inv_data[hs1]["HIpols"])-1))
            mgeodict = geoid.ArcDirect(inv_data[hs1]["HILat"],inv_data[hs1]["HILon"],inv_data[hs1]["HI_start_azi"]+midpoint_dis,-inv_data[hs1]["HIDis"]) #modeled geographic point for age
            midpoint = [mgeodict["lat2"],(360+mgeodict["lon2"])%360]
            midpoint_azi = mgeodict["azi2"]
            dage = np.sqrt((inv_data[hs1]["HIsds"][0]**2)*(midpoint_age-inv_data[hs1]["HI_mean_age"])**2 + inv_data[hs1]["HIsds"][1]**2)
        else:  ##########Emperor
            hs1_b = np.sqrt(2)*(33./111.113)*np.sqrt(inv_data[hs1]["EMErr"]/(len(data.index)-6))
            if len(inv_data[hs1]["EMpols"])==2:
                midpoint_dis = ((midpoint_age-inv_data[hs1]["EMpols"][1])/inv_data[hs1]["EMpols"][0] - inv_data[hs1]["EM_Start_Dis"])/np.sin(np.deg2rad(inv_data[hs1]["EMDis"]))
                hs1_a = np.sqrt(2*(((inv_data[hs1]["EMsds"][1]**2)*(1/inv_data[hs1]["EMpols"][0])**2) + ((inv_data[hs1]["EMsds"][0]**2)*(-(midpoint_age-inv_data[hs1]["EM_mean_age"]-inv_data[hs1]["EMpols"][1])/(inv_data[hs1]["EMpols"][0]**2))**2)) + hs1_b**2)

#                center_pols,center_sds = utl.polyrecenter(inv_data[hs1]["EMpols"],[midpoint_age],np.ones(1)*inv_data[hs1]["EMAgeSd"],center=inv_data[hs1]["EM_mean_dis"])
#                center_pols,center_sds = inv_data[hs1]["EMpols"],inv_data[hs1]["EMsds"]
#                hi_err_env = utl.polyerr(center_pols,center_sds,[midpoint_age],xerr=None)
#                hs1_a = (1/inv_data[hs1]["EMpols"][0])*hi_err_env[0]
#                hs1_a = np.sqrt(2)*((1/inv_data[hs1]["EMpols"][0])*inv_data[hs1]["HIAgeSd"])
            elif len(inv_data[hs1]["EMpols"])==3:
                midpoint_dis = ((((-inv_data[hs1]["EMpols"][1]+np.sqrt(inv_data[hs1]["EMpols"][1]**2 - 4*inv_data[hs1]["EMpols"][0]*(inv_data[hs1]["EMpols"][2]-midpoint_age)))/(2*inv_data[hs1]["EMpols"][0]))) - inv_data[hs1]["EM_Start_Dis"])/np.sin(np.deg2rad(inv_data[hs1]["EMDis"]))
                psi = np.sqrt(inv_data[hs1]["EMpols"][1]**2 - 4*inv_data[hs1]["EMpols"][0]*(inv_data[hs1]["EMpols"][2]-midpoint_age))
                d_dc = 1/psi
                d_db = (inv_data[hs1]["EMpols"][1]/psi-1)/(2*inv_data[hs1]["EMpols"][0])
                d_da = (psi-inv_data[hs1]["EMpols"][1])/(2*inv_data[hs1]["EMpols"][0]**2) - (inv_data[hs1]["EMpols"][2]-midpoint_age)/(inv_data[hs1]["EMpols"][0]*psi)
                hs1_a = np.sqrt(2*len(x)*(((center_sds[2]**2)*(d_dc**2)) + ((center_sds[1]**2)*(d_db**2)) + ((center_sds[0]**2)*(d_da**2))) + hs1_b**2)
            else: raise ValueError("Degree %d inverse not programmed yet"%(len(inv_data[hs1]["EMpols"])-1))
            mgeodict = geoid.ArcDirect(inv_data[hs1]["EMLat"],inv_data[hs1]["EMLon"],inv_data[hs1]["EM_start_azi"]+midpoint_dis,-inv_data[hs1]["EMDis"]) #modeled geograpEMc point for age
            midpoint = [mgeodict["lat2"],(360+mgeodict["lon2"])%360]
            midpoint_azi = mgeodict["azi2"]
            dage = np.sqrt((inv_data[hs1]["EMsds"][0]**2)*(midpoint_age-inv_data[hs1]["EM_mean_age"])**2 + inv_data[hs1]["EMsds"][1]**2)

        hs_df.at[idx,"Age"] = midpoint_age
        hs_df.at[idx,"Azi"] = midpoint_azi-90
        hs_df.at[idx,"Comment"] = ""
        hs_df.at[idx,"Lat"] = midpoint[0]
        hs_df.at[idx,"Lon"] = midpoint[1]
        hs_df.at[idx,"Loc"] = "%.0f Ma Linear Age Model"%midpoint_age
        hs_df.at[idx,"MajSE"] = hs1_a
        hs_df.at[idx,"MinSE"] = hs1_b
        hs_df.at[idx,"Plate"] = "PA"
        hs_df.at[idx,"Ref"] = "Gaastra et al. 2020"
        hs_df.at[idx,"dAge"] = dage

        print("%.0f Ma:"%midpoint_age,*midpoint,midpoint_azi-90,hs1_a,hs1_b)
        bm = psk.plot_pole(midpoint[1],midpoint[0],midpoint_azi-90,hs1_a,hs1_b,edgecolors="k",facecolors=color,color=color,marker="s",m=bm,zorder=10000,s=15)

    #Plot Bend Figures
    all_lons,all_lats,all_grav = pg.get_sandwell(bend_extent,decimation,sandwell_files_path="../PySkew/raw_data/gravity/Sandwell/*.tiff")

    print("Plotting Gravity")
    start_time = time()
    print(all_lons.shape,all_lats.shape,all_grav.shape)
#    potental cmaps: cividis
    fcm = bm.contourf(all_lons, all_lats, all_grav, cmap="Blues_r", alpha=.75, transform=ccrs.PlateCarree(), zorder=0)
    print("Runtime: ",time()-start_time)

    print(bend_extent)
    bm.set_extent(bend_extent, ccrs.PlateCarree())

    print("HICIRC",inv_data[hs1]["HILat"],inv_data[hs1]["HILon"],inv_data[hs1]["HIDis"],geoid.Inverse(inv_data[hs1]["HILat"],inv_data[hs1]["HILon"],*midpoint)["a12"],geoid.Inverse(inv_data[hs1]["HILat"],inv_data[hs1]["HILon"],*midpoint)["azi2"])
    bm = psk.plot_small_circle(inv_data[hs1]["HILon"],inv_data[hs1]["HILat"],inv_data[hs1]["HIDis"],m=bm,color="k",linewidth=1.5,zorder=2,geoid=geoid)
    bm = psk.plot_small_circle(inv_data[hs1]["HILon"],inv_data[hs1]["HILat"],inv_data[hs1]["HIDis"],m=bm,color=hi_color,linewidth=1,zorder=3,geoid=geoid)
    bm = psk.plot_small_circle(inv_data[hs1]["EMLon"],inv_data[hs1]["EMLat"],inv_data[hs1]["EMDis"],m=bm,color="k",linewidth=1.5,zorder=2,geoid=geoid)
    bm = psk.plot_small_circle(inv_data[hs1]["EMLon"],inv_data[hs1]["EMLat"],inv_data[hs1]["EMDis"],m=bm,color=em_color,linewidth=1,zorder=3,geoid=geoid)

    if "Louisville" in hs1: title = "Louisville"
    else: title = hs1
    if k==0: bm.set_title("Hawaiian Age")
    bm.set_ylabel(title)
    bax_pos += 1

    print("Writing: %s to %s"%(hs1,fout))
    f = append_df_to_excel(fout, hs_df, sheet_name=hs1, index=False, startrow=0)
    f.save(); f.close()

print("Saving: %s"%(os.path.join(os.path.dirname(fin),os.path.basename(fin).split(".")[0]+"_recdata.png")))
bfig.savefig(os.path.join(os.path.dirname(fin),os.path.basename(fin).split(".")[0]+"_recdata.png"))
print("Saving: %s"%(os.path.join(os.path.dirname(fin),os.path.basename(fin).split(".")[0]+"_recdata.pdf")))
#bfig.savefig(os.path.join(os.path.dirname(fin),os.path.basename(fin).split(".")[0]+"_recdata.pdf"))

